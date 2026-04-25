"""
Build Star-Energy-CEMS-Cost-Tracker.xlsx — business-standard workbook.
Run: python3 build-xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from pathlib import Path

# ─────────────────────── Style palette ───────────────────────
NAVY = "1F3A5F"
TEAL = "2A7F62"
AMBER = "D98E04"
SLATE = "475569"
LIGHT = "F1F5F9"
MID = "CBD5E1"
RED_SOFT = "FCA5A5"
GREEN_SOFT = "86EFAC"

thin = Side(border_style="thin", color="94A3B8")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_fill(): return PatternFill("solid", fgColor=NAVY)
def subtotal_fill(): return PatternFill("solid", fgColor=LIGHT)
def total_fill(): return PatternFill("solid", fgColor=MID)

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=NAVY)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color=NAVY)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

FMT_USD = '"$"#,##0.00'
FMT_USD_WHOLE = '"$"#,##0'
FMT_HRS = '#,##0.0'
FMT_PCT = '0.0%'

def set_header(ws, row, headers):
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = header_fill()
        c.alignment = ALIGN_CENTER
        c.border = border_all
    ws.row_dimensions[row].height = 28

def style_row(ws, row, n_cols, font=None, fill=None):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        if font: c.font = font
        if fill: c.fill = fill
        c.border = border_all
        if not c.alignment or c.alignment.horizontal is None:
            c.alignment = ALIGN_LEFT

def autowidth(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, text, span_cols):
    ws.cell(row=1, column=1, value=text).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span_cols)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

wb = Workbook()
wb.remove(wb.active)

# ═══════════════════════════════════════════════════════════════
# Sheet 2: Labor (build first so dashboard can reference)
# ═══════════════════════════════════════════════════════════════
ws_labor = wb.create_sheet("Labor")
ws_labor.sheet_properties.tabColor = TEAL
title_row(ws_labor, "Labor — MVP Stories (2-Terminal Parallel Plan)", 13)

labor_headers = ["Epic", "Story ID", "Story Name", "Terminal",
                 "Claude Work Hrs", "Your Rate (USD)", "Est Cost (USD)",
                 "Actual Hrs", "Actual Cost (USD)", "Variance Hrs",
                 "Variance Cost", "Status", "Notes"]
set_header(ws_labor, 3, labor_headers)

# (Epic, ID, Name, Terminal, Hrs, Status, Notes)
rows = [
    ("Epic 0", "0-1", "Turborepo monorepo + shared package scaffold", "Done", 0, "done", "Already shipped"),
    ("Epic 0", "0-2", "Azure infrastructure provisioning", "Done", 0, "done", "Already shipped"),
    ("Epic 0", "0-3", "Database schema + RLS foundation", "Seq (A)", 12, "in-progress", "Sequential blocker — A alone"),
    ("Epic 0", "0-4", "Node.js API foundation + BullMQ", "A", 12, "backlog", ""),
    ("Epic 0", "0-5", "Python calculation service scaffold", "DEFERRED", 0, "deferred", "Consolidate calcs into Node for MVP"),
    ("Epic 0", "0-6", "CI/CD pipeline", "A", 6, "backlog", "Basic GitHub Actions"),
    ("Epic 0", "0-7", "Design tokens + minimal component library", "B", 4, "backlog", "Tailwind defaults + 3-4 components"),
    ("Epic 0", "0-8", "Accessibility + testing infrastructure", "DEFERRED", 0, "deferred", "V2 polish"),
    ("Epic 1", "1-1", "Login API + JWT issuance", "A", 8, "backlog", ""),
    ("Epic 1", "1-1b", "Login UI screen", "B", 6, "backlog", ""),
    ("Epic 1", "1-2", "Role-based access API guards", "A", 4, "backlog", "Auditor + admin roles only"),
    ("Epic 1", "1-2b", "Route guards — frontend", "B", 4, "backlog", ""),
    ("Epic 1", "1-3", "Admin user management — auditors", "A+B", 8, "backlog", "Simple CRUD + list"),
    ("Epic 1", "1-4", "Admin client management", "DEFERRED", 0, "deferred", "No client accounts in MVP"),
    ("Epic 2", "2-1", "Store reference API", "A", 6, "backlog", ""),
    ("Epic 2", "2-1b", "Store selector screen", "B", 6, "backlog", ""),
    ("Epic 2", "2-2", "Auto-fill + draft creation API", "A", 8, "backlog", ""),
    ("Epic 2", "2-2b", "Auto-fill UI", "B", 4, "backlog", ""),
    ("Epic 2", "2-3", "Auto-save API endpoints", "A", 6, "backlog", ""),
    ("Epic 2", "2-3b", "Auto-save frontend hook", "B", 4, "backlog", ""),
    ("Epic 2", "2-4", "Section overview + nav shell", "B", 8, "backlog", ""),
    ("Epic 3", "3-1", "Audit breadcrumb + machine room screen", "B", 8, "backlog", ""),
    ("Epic 3", "3-2", "Rack entry + duplication", "DEFERRED", 0, "deferred", "V2 — single rack in MVP"),
    ("Epic 3", "3-3a", "Compressor + regression API", "A", 10, "backlog", "Regression DB lookup"),
    ("Epic 3", "3-3b", "Compressor screens UI", "B", 12, "backlog", ""),
    ("Epic 3", "3-4", "Condenser + display case UI + API", "A+B", 14, "backlog", "Walk-in coolers deferred"),
    ("Epic 3", "3-5", "Controller systems", "DEFERRED", 0, "deferred", "V2"),
    ("Epic 3", "3-6", "Multi-auditor locking", "DEFERRED", 0, "deferred", "Single-auditor MVP"),
    ("Epic 4", "4-1", "Camera + PhotoCaptureField", "B", 12, "backlog", ""),
    ("Epic 4", "4-2", "Auto-tag + Azure Blob upload API", "A", 10, "backlog", ""),
    ("Epic 4", "4-3", "Additional photos + comments", "B", 6, "backlog", ""),
    ("Epic 4", "4-4", "Upload queue retry", "DEFERRED", 0, "deferred", "Fail-loud MVP"),
    ("Epic 5", "5-1", "General section screens (minimal)", "B", 6, "backlog", ""),
    ("Epic 5", "5-2", "HVAC section", "DEFERRED", 0, "deferred", "V2"),
    ("Epic 5", "5-3", "Lighting + envelope", "DEFERRED", 0, "deferred", "V2"),
    ("Epic 5", "5-4", "Review + submit screen", "B", 8, "backlog", ""),
    ("Epic 5", "5-5", "Submission confirmation + email", "A+B", 6, "backlog", "SendGrid free tier"),
    ("Epic 6", "6-1", "Store reference seed script", "A", 3, "backlog", "SQL seed — no UI"),
    ("Epic 6", "6-2", "Compressor regression seed script", "A", 3, "backlog", "SQL seed — no UI"),
    ("Epic 6", "6-3", "Weather data retrieval", "DEFERRED", 0, "deferred", "Not needed without baseline"),
    ("Epic 7", "7-1", "Audit queue API + list UI (simple)", "A+B", 6, "backlog", "No filters; simple list"),
    ("Epic 7", "7-2", "SLA timer + detail sheet", "DEFERRED", 0, "deferred", "V2"),
    ("Epic 7", "7-3", "State machine transitions", "DEFERRED", 0, "deferred", "Simple transitions in MVP"),
    ("Epic 7", "7-4", "Photo lock enforcement", "DEFERRED", 0, "deferred", "V2"),
    ("Epic 8", "8-1", "ECM refrigeration calc", "A", 24, "backlog", "Core calc — Node for MVP"),
    ("Epic 8", "8-2", "Energy baseline regression", "DEFERRED", 0, "deferred", "V2"),
    ("Epic 8", "8-3", "LLM anomaly flagging", "DEFERRED", 0, "deferred", "V2 — biggest deferred feature"),
    ("Epic 8", "8-4", "Admin flag override", "DEFERRED", 0, "deferred", "Needs 8-3"),
    ("Epic 8", "8-5", "Calc inputs/outputs view", "B", 6, "backlog", "Basic read-only"),
    ("Epic 9", "9-1", "Puppeteer PDF renderer", "A", 12, "backlog", "Self-contained HTML bundle"),
    ("Epic 9", "9-2a", "Admin approve + publish API", "A", 4, "backlog", ""),
    ("Epic 9", "9-2b", "Admin approve UI", "B", 4, "backlog", ""),
    ("Epic 9", "9-3", "Report immutability / versioning", "DEFERRED", 0, "deferred", "V2"),
    ("Epic 9", "9-4", "Client email on publish", "A", 4, "backlog", "SendGrid"),
    ("Epic 9", "9-5", "Report regeneration", "DEFERRED", 0, "deferred", "V2"),
    ("Epic 10", "10-1", "Client portal dashboard", "DEFERRED", 0, "deferred", "Email PDF instead"),
    ("Epic 10", "10-2", "Client portal report view", "DEFERRED", 0, "deferred", "Email PDF instead"),
    ("Epic 10", "10-3", "Client portal ECM summary", "DEFERRED", 0, "deferred", "V2"),
    ("Epic 10", "10-4", "Client portal notifications", "DEFERRED", 0, "deferred", "V2"),
]

RATE = 20
start_row = 4
for i, (epic, sid, name, terminal, hrs, status, notes) in enumerate(rows):
    r = start_row + i
    ws_labor.cell(row=r, column=1, value=epic).font = BODY_FONT
    ws_labor.cell(row=r, column=2, value=sid).font = BODY_FONT
    ws_labor.cell(row=r, column=3, value=name).font = BODY_FONT
    ws_labor.cell(row=r, column=4, value=terminal).font = BODY_FONT
    c5 = ws_labor.cell(row=r, column=5, value=hrs); c5.font = BODY_FONT; c5.number_format = FMT_HRS
    c6 = ws_labor.cell(row=r, column=6, value=RATE); c6.font = BODY_FONT; c6.number_format = FMT_USD_WHOLE
    c7 = ws_labor.cell(row=r, column=7, value=f"=E{r}*F{r}"); c7.number_format = FMT_USD
    c8 = ws_labor.cell(row=r, column=8); c8.number_format = FMT_HRS
    c9 = ws_labor.cell(row=r, column=9, value=f"=H{r}*F{r}"); c9.number_format = FMT_USD
    c10 = ws_labor.cell(row=r, column=10, value=f"=H{r}-E{r}"); c10.number_format = FMT_HRS
    c11 = ws_labor.cell(row=r, column=11, value=f"=I{r}-G{r}"); c11.number_format = FMT_USD
    ws_labor.cell(row=r, column=12, value=status).font = BODY_FONT
    ws_labor.cell(row=r, column=13, value=notes).font = BODY_FONT
    for col in range(1, 14):
        ws_labor.cell(row=r, column=col).border = border_all
    if terminal == "DEFERRED":
        for col in range(1, 14):
            ws_labor.cell(row=r, column=col).fill = PatternFill("solid", fgColor="F3F4F6")
            ws_labor.cell(row=r, column=col).font = Font(name="Calibri", size=10, italic=True, color="9CA3AF")

end_data = start_row + len(rows) - 1  # last data row

# Subtotals
r_sub = end_data + 2
ws_labor.cell(row=r_sub, column=3, value="MVP In-Scope Claude Work Hours")
ws_labor.cell(row=r_sub, column=5, value=f"=SUM(E{start_row}:E{end_data})").number_format = FMT_HRS
ws_labor.cell(row=r_sub, column=7, value=f"=SUM(G{start_row}:G{end_data})").number_format = FMT_USD
ws_labor.cell(row=r_sub, column=8, value=f"=SUM(H{start_row}:H{end_data})").number_format = FMT_HRS
ws_labor.cell(row=r_sub, column=9, value=f"=SUM(I{start_row}:I{end_data})").number_format = FMT_USD
ws_labor.cell(row=r_sub, column=10, value=f"=SUM(J{start_row}:J{end_data})").number_format = FMT_HRS
ws_labor.cell(row=r_sub, column=11, value=f"=SUM(K{start_row}:K{end_data})").number_format = FMT_USD
style_row(ws_labor, r_sub, 13, font=BOLD_FONT, fill=subtotal_fill())

r_a = r_sub + 1
ws_labor.cell(row=r_a, column=3, value="Terminal A sequential hours")
ws_labor.cell(row=r_a, column=5,
    value=f'=SUMIFS(E{start_row}:E{end_data},D{start_row}:D{end_data},"A")'
          f'+SUMIFS(E{start_row}:E{end_data},D{start_row}:D{end_data},"Seq (A)")'
          f'+SUMIFS(E{start_row}:E{end_data},D{start_row}:D{end_data},"A+B")/2').number_format = FMT_HRS
style_row(ws_labor, r_a, 13, font=BODY_FONT, fill=subtotal_fill())

r_b = r_sub + 2
ws_labor.cell(row=r_b, column=3, value="Terminal B sequential hours")
ws_labor.cell(row=r_b, column=5,
    value=f'=SUMIFS(E{start_row}:E{end_data},D{start_row}:D{end_data},"B")'
          f'+SUMIFS(E{start_row}:E{end_data},D{start_row}:D{end_data},"A+B")/2').number_format = FMT_HRS
style_row(ws_labor, r_b, 13, font=BODY_FONT, fill=subtotal_fill())

r_cal = r_sub + 3
ws_labor.cell(row=r_cal, column=3, value="Calendar time (slower terminal = critical path)")
ws_labor.cell(row=r_cal, column=5, value=f"=MAX(E{r_a},E{r_b})").number_format = FMT_HRS
style_row(ws_labor, r_cal, 13, font=BODY_FONT, fill=subtotal_fill())

r_eff = r_sub + 4
ws_labor.cell(row=r_eff, column=3, value="Your effective hours (÷ 1.4 parallel efficiency)")
ws_labor.cell(row=r_eff, column=5, value=f"=E{r_sub}/1.4").number_format = FMT_HRS
ws_labor.cell(row=r_eff, column=6, value=RATE).number_format = FMT_USD_WHOLE
ws_labor.cell(row=r_eff, column=7, value=f"=E{r_eff}*F{r_eff}").number_format = FMT_USD
style_row(ws_labor, r_eff, 13, font=TOTAL_FONT, fill=total_fill())

r_tot = r_sub + 5
ws_labor.cell(row=r_tot, column=3, value="MVP LABOR COST @ your rate")
ws_labor.cell(row=r_tot, column=7, value=f"=G{r_eff}").number_format = FMT_USD
style_row(ws_labor, r_tot, 13, font=TOTAL_FONT, fill=PatternFill("solid", fgColor=MID))

# Named ranges for dashboard
wb.defined_names["LABOR_HOURS_TOTAL"] = DefinedName("LABOR_HOURS_TOTAL", attr_text=f"Labor!$E${r_sub}")
wb.defined_names["LABOR_EFFECTIVE_HOURS"] = DefinedName("LABOR_EFFECTIVE_HOURS", attr_text=f"Labor!$E${r_eff}")
wb.defined_names["LABOR_COST"] = DefinedName("LABOR_COST", attr_text=f"Labor!$G${r_eff}")

autowidth(ws_labor, [10, 8, 42, 10, 12, 12, 14, 12, 14, 12, 14, 13, 32])
ws_labor.freeze_panes = "A4"

# Conditional formatting on variance cost column
ws_labor.conditional_formatting.add(f"K{start_row}:K{end_data}",
    CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill("solid", fgColor=RED_SOFT)))
ws_labor.conditional_formatting.add(f"K{start_row}:K{end_data}",
    CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill("solid", fgColor=GREEN_SOFT)))

# ═══════════════════════════════════════════════════════════════
# Sheet 3: Subscriptions
# ═══════════════════════════════════════════════════════════════
ws_subs = wb.create_sheet("Subscriptions")
ws_subs.sheet_properties.tabColor = AMBER
title_row(ws_subs, "Subscriptions & Dev Tools (2-month MVP build)", 10)
subs_headers = ["Category", "Item", "Plan", "Monthly (USD)", "Months", "Est Total (USD)",
                "Actual Total (USD)", "Variance (USD)", "Vendor", "Required?"]
set_header(ws_subs, 3, subs_headers)
subs = [
    ("Claude", "Claude Code Max 20x", "Max 20x", 200, 2, "Anthropic", "Required"),
    ("Source Control", "GitHub Free", "Free", 0, 2, "GitHub", "Required"),
    ("Design", "Figma Free", "Free", 0, 2, "Figma", "Optional"),
    ("Password Mgmt", "1Password (optional)", "Individual", 3, 2, "1Password", "Optional"),
    ("Error Tracking", "Sentry Developer", "Free", 0, 2, "Sentry", "Recommended"),
    ("Project Mgmt", "Notion / Linear", "Free", 0, 2, "Notion", "Optional"),
    ("Diagrams", "Excalidraw", "Free", 0, 2, "Excalidraw", "Optional"),
]
r = 4
for cat, item, plan, monthly, months, vendor, req in subs:
    ws_subs.cell(row=r, column=1, value=cat).font = BODY_FONT
    ws_subs.cell(row=r, column=2, value=item).font = BODY_FONT
    ws_subs.cell(row=r, column=3, value=plan).font = BODY_FONT
    ws_subs.cell(row=r, column=4, value=monthly).number_format = FMT_USD
    ws_subs.cell(row=r, column=5, value=months).number_format = FMT_HRS
    ws_subs.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = FMT_USD
    ws_subs.cell(row=r, column=7).number_format = FMT_USD
    ws_subs.cell(row=r, column=8, value=f"=G{r}-F{r}").number_format = FMT_USD
    ws_subs.cell(row=r, column=9, value=vendor).font = BODY_FONT
    ws_subs.cell(row=r, column=10, value=req).font = BODY_FONT
    for col in range(1, 11):
        ws_subs.cell(row=r, column=col).border = border_all
    r += 1
r_total_subs = r + 1
ws_subs.cell(row=r_total_subs, column=2, value="TOTAL (required only — Claude Max)")
ws_subs.cell(row=r_total_subs, column=6, value="=F4").number_format = FMT_USD
ws_subs.cell(row=r_total_subs, column=8, value="=H4").number_format = FMT_USD
style_row(ws_subs, r_total_subs, 10, font=TOTAL_FONT, fill=total_fill())
ws_subs.cell(row=r_total_subs+1, column=2, value="TOTAL (all lines)")
ws_subs.cell(row=r_total_subs+1, column=6, value=f"=SUM(F4:F{r-1})").number_format = FMT_USD
ws_subs.cell(row=r_total_subs+1, column=8, value=f"=SUM(H4:H{r-1})").number_format = FMT_USD
style_row(ws_subs, r_total_subs+1, 10, font=BOLD_FONT, fill=subtotal_fill())

wb.defined_names["SUBS_REQUIRED_TOTAL"] = DefinedName("SUBS_REQUIRED_TOTAL", attr_text=f"Subscriptions!$F${r_total_subs}")
wb.defined_names["SUBS_ALL_TOTAL"] = DefinedName("SUBS_ALL_TOTAL", attr_text=f"Subscriptions!$F${r_total_subs+1}")

autowidth(ws_subs, [15, 28, 20, 14, 10, 15, 16, 14, 14, 13])
ws_subs.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════
# Sheet 4: Infrastructure
# ═══════════════════════════════════════════════════════════════
ws_infra = wb.create_sheet("Infrastructure")
ws_infra.sheet_properties.tabColor = AMBER
title_row(ws_infra, "Azure Infrastructure + Third-Party (MVP build period)", 8)
infra_headers = ["Environment", "Service", "Tier / SKU", "Monthly (USD)", "Months",
                 "Est Total (USD)", "Actual (USD)", "Variance (USD)"]
set_header(ws_infra, 3, infra_headers)
infra_sections = [
    ("Dev", [
        ("Azure SQL Database", "Basic (5 DTU)", 5, 2),
        ("Azure App Service (API + PDF combined)", "B1 Linux", 13, 2),
        ("Azure Static Web Apps x2", "Free tier", 0, 2),
        ("Azure Blob Storage", "Hot <5GB", 2, 2),
        ("Azure Cache for Redis", "Basic C0", 17, 2),
        ("Application Insights", "Free tier", 0, 2),
        ("Azure Key Vault", "Standard", 3, 2),
    ]),
    ("Prod", [
        ("Azure SQL Database", "S0 (10 DTU)", 15, 1),
        ("Azure App Service (API + PDF)", "B2 Linux", 55, 1),
        ("Azure Static Web Apps x2", "Standard", 9, 1),
        ("Azure Blob Storage", "Hot + 10GB", 5, 1),
        ("Azure Cache for Redis", "Basic C0", 17, 1),
        ("Application Insights", "Pay-as-you-go", 10, 1),
        ("Azure Key Vault", "Standard", 3, 1),
        ("Bandwidth / Egress", "Canada Central", 5, 1),
    ]),
    ("3rd-party", [
        ("SendGrid", "Free (100/day)", 0, 2),
        ("Weather API", "NOT NEEDED (deferred)", 0, 2),
        ("Domain (amortized)", ".com", 1, 2),
        ("Managed SSL", "Azure (free)", 0, 2),
    ]),
]

r = 4
section_totals = {}
for env, items in infra_sections:
    sec_start = r
    for svc, tier, monthly, months in items:
        ws_infra.cell(row=r, column=1, value=env).font = BODY_FONT
        ws_infra.cell(row=r, column=2, value=svc).font = BODY_FONT
        ws_infra.cell(row=r, column=3, value=tier).font = BODY_FONT
        ws_infra.cell(row=r, column=4, value=monthly).number_format = FMT_USD
        ws_infra.cell(row=r, column=5, value=months).number_format = FMT_HRS
        ws_infra.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = FMT_USD
        ws_infra.cell(row=r, column=7).number_format = FMT_USD
        ws_infra.cell(row=r, column=8, value=f"=G{r}-F{r}").number_format = FMT_USD
        for col in range(1, 9):
            ws_infra.cell(row=r, column=col).border = border_all
        r += 1
    # subtotal
    sec_end = r - 1
    ws_infra.cell(row=r, column=2, value=f"{env} SUBTOTAL")
    ws_infra.cell(row=r, column=6, value=f"=SUM(F{sec_start}:F{sec_end})").number_format = FMT_USD
    ws_infra.cell(row=r, column=7, value=f"=SUM(G{sec_start}:G{sec_end})").number_format = FMT_USD
    ws_infra.cell(row=r, column=8, value=f"=SUM(H{sec_start}:H{sec_end})").number_format = FMT_USD
    style_row(ws_infra, r, 8, font=BOLD_FONT, fill=subtotal_fill())
    section_totals[env] = r
    r += 2

r_grand = r
ws_infra.cell(row=r_grand, column=2, value="GRAND TOTAL (Dev + Prod launch + 3rd-party)")
ws_infra.cell(row=r_grand, column=6,
    value=f"=F{section_totals['Dev']}+F{section_totals['Prod']}+F{section_totals['3rd-party']}").number_format = FMT_USD
ws_infra.cell(row=r_grand, column=7,
    value=f"=G{section_totals['Dev']}+G{section_totals['Prod']}+G{section_totals['3rd-party']}").number_format = FMT_USD
ws_infra.cell(row=r_grand, column=8, value=f"=G{r_grand}-F{r_grand}").number_format = FMT_USD
style_row(ws_infra, r_grand, 8, font=TOTAL_FONT, fill=total_fill())

wb.defined_names["INFRA_DEV_TOTAL"] = DefinedName("INFRA_DEV_TOTAL", attr_text=f"Infrastructure!$F${section_totals['Dev']}")
wb.defined_names["INFRA_PROD_TOTAL"] = DefinedName("INFRA_PROD_TOTAL", attr_text=f"Infrastructure!$F${section_totals['Prod']}")
wb.defined_names["INFRA_3P_TOTAL"] = DefinedName("INFRA_3P_TOTAL", attr_text=f"Infrastructure!$F${section_totals['3rd-party']}")
wb.defined_names["INFRA_GRAND_TOTAL"] = DefinedName("INFRA_GRAND_TOTAL", attr_text=f"Infrastructure!$F${r_grand}")

autowidth(ws_infra, [12, 36, 22, 14, 10, 15, 14, 14])
ws_infra.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════
# Sheet 5: One-Time Costs
# ═══════════════════════════════════════════════════════════════
ws_one = wb.create_sheet("One-Time")
ws_one.sheet_properties.tabColor = AMBER
title_row(ws_one, "One-Time Costs (MVP launch)", 7)
one_headers = ["Category", "Item", "Est (USD)", "Actual (USD)", "Variance (USD)", "Required?", "Notes"]
set_header(ws_one, 3, one_headers)
one_rows = [
    ("Branding", "Logo + brand kit", 0, "Skip", "Use text logo / existing Star Energy branding"),
    ("Branding", "Illustrations / icons", 0, "Skip", "Free Heroicons / Lucide"),
    ("Legal", "T&Cs + Privacy (Termly template)", 50, "Recommended", "PIPEDA basics"),
    ("Legal", "Lawyer review", 0, "Skip", "Defer to post-MVP"),
    ("Compliance", "Security audit / pentest", 0, "Skip", "Defer to pre-client-rollout"),
    ("Compliance", "SOC2 readiness", 0, "Skip", "V2+"),
    ("Domain", "Primary domain (.com)", 15, "Required", "Annual"),
    ("Domain", "Domain privacy", 0, "Optional", "Often free"),
    ("SSL", "Managed SSL", 0, "Required", "Azure App Service managed cert (free)"),
    ("Setup", "Backup / DR config", 0, "Required", "SQL 7-day PITR included"),
    ("Training", "Auditor onboarding (doc + Loom)", 50, "Recommended", ""),
    ("Training", "Client rollout walkthrough", 0, "Skip", "Defer until first client"),
    ("Launch", "Marketing / landing page", 0, "Skip", "Not needed for internal tool"),
    ("Contingency", "Unknown unknowns buffer", 50, "Recommended", "~15% of one-time"),
]
r = 4
for cat, item, est, req, notes in one_rows:
    ws_one.cell(row=r, column=1, value=cat).font = BODY_FONT
    ws_one.cell(row=r, column=2, value=item).font = BODY_FONT
    ws_one.cell(row=r, column=3, value=est).number_format = FMT_USD
    ws_one.cell(row=r, column=4).number_format = FMT_USD
    ws_one.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
    ws_one.cell(row=r, column=6, value=req).font = BODY_FONT
    ws_one.cell(row=r, column=7, value=notes).font = BODY_FONT
    for col in range(1, 8):
        ws_one.cell(row=r, column=col).border = border_all
    r += 1
r_one_total = r
ws_one.cell(row=r_one_total, column=2, value="TOTAL")
ws_one.cell(row=r_one_total, column=3, value=f"=SUM(C4:C{r-1})").number_format = FMT_USD
ws_one.cell(row=r_one_total, column=4, value=f"=SUM(D4:D{r-1})").number_format = FMT_USD
ws_one.cell(row=r_one_total, column=5, value=f"=SUM(E4:E{r-1})").number_format = FMT_USD
style_row(ws_one, r_one_total, 7, font=TOTAL_FONT, fill=total_fill())
wb.defined_names["ONETIME_TOTAL"] = DefinedName("ONETIME_TOTAL", attr_text=f"'One-Time'!$C${r_one_total}")

autowidth(ws_one, [15, 34, 13, 13, 14, 15, 40])
ws_one.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════
# Sheet 6: Production Run-Rate
# ═══════════════════════════════════════════════════════════════
ws_run = wb.create_sheet("Production Run-Rate")
ws_run.sheet_properties.tabColor = "7C3AED"
title_row(ws_run, "Year 1 Production Run-Rate (post-launch)", 5)
run_headers = ["Category", "Line Item", "Monthly (USD)", "Annual (USD)", "Notes"]
set_header(ws_run, 3, run_headers)
run_rows = [
    ("Infrastructure", "Azure SQL (S0 → S1 after beta)", 15, "Upgrade when >100 audits/mo"),
    ("Infrastructure", "Azure App Service B2 (API + PDF)", 55, "Split only under memory pressure"),
    ("Infrastructure", "Azure Static Web Apps x2 Standard", 18, "Custom domain support"),
    ("Infrastructure", "Azure Blob Storage", 10, "Hot → Cool lifecycle after 90d"),
    ("Infrastructure", "Azure Cache for Redis (C0)", 17, "Upgrade to C1 if queue backs up"),
    ("Infrastructure", "Application Insights", 10, "Low initial traffic"),
    ("Infrastructure", "Key Vault + Monitor + bandwidth", 15, "Aggregated"),
]
r = 4
for cat, item, monthly, notes in run_rows:
    ws_run.cell(row=r, column=1, value=cat).font = BODY_FONT
    ws_run.cell(row=r, column=2, value=item).font = BODY_FONT
    ws_run.cell(row=r, column=3, value=monthly).number_format = FMT_USD
    ws_run.cell(row=r, column=4, value=f"=C{r}*12").number_format = FMT_USD
    ws_run.cell(row=r, column=5, value=notes).font = BODY_FONT
    for col in range(1, 6):
        ws_run.cell(row=r, column=col).border = border_all
    r += 1
r_infra_sub = r
ws_run.cell(row=r_infra_sub, column=2, value="INFRA SUBTOTAL")
ws_run.cell(row=r_infra_sub, column=3, value=f"=SUM(C4:C{r-1})").number_format = FMT_USD
ws_run.cell(row=r_infra_sub, column=4, value=f"=SUM(D4:D{r-1})").number_format = FMT_USD
style_row(ws_run, r_infra_sub, 5, font=BOLD_FONT, fill=subtotal_fill())

r = r_infra_sub + 2
ops_rows = [
    ("3rd-party", "SendGrid Essentials (post-beta)", 20, "When >100 emails/day"),
    ("3rd-party", "Domain + managed SSL", 1, "$15/yr amortized"),
    ("3rd-party", "Sentry Developer (free tier)", 0, "Until >5k errors/mo"),
    ("Backup", "Extended backup retention", 5, "Beyond default 7 days"),
]
ops_start = r
for cat, item, monthly, notes in ops_rows:
    ws_run.cell(row=r, column=1, value=cat).font = BODY_FONT
    ws_run.cell(row=r, column=2, value=item).font = BODY_FONT
    ws_run.cell(row=r, column=3, value=monthly).number_format = FMT_USD
    ws_run.cell(row=r, column=4, value=f"=C{r}*12").number_format = FMT_USD
    ws_run.cell(row=r, column=5, value=notes).font = BODY_FONT
    for col in range(1, 6):
        ws_run.cell(row=r, column=col).border = border_all
    r += 1
r_ops_sub = r
ws_run.cell(row=r_ops_sub, column=2, value="OPS SUBTOTAL")
ws_run.cell(row=r_ops_sub, column=3, value=f"=SUM(C{ops_start}:C{r-1})").number_format = FMT_USD
ws_run.cell(row=r_ops_sub, column=4, value=f"=SUM(D{ops_start}:D{r-1})").number_format = FMT_USD
style_row(ws_run, r_ops_sub, 5, font=BOLD_FONT, fill=subtotal_fill())

r_contingency = r_ops_sub + 1
ws_run.cell(row=r_contingency, column=2, value="Contingency (10%)")
ws_run.cell(row=r_contingency, column=3, value=f"=(C{r_infra_sub}+C{r_ops_sub})*0.1").number_format = FMT_USD
ws_run.cell(row=r_contingency, column=4, value=f"=C{r_contingency}*12").number_format = FMT_USD
style_row(ws_run, r_contingency, 5, font=BODY_FONT, fill=subtotal_fill())

r_run_total = r_contingency + 1
ws_run.cell(row=r_run_total, column=2, value="MVP MONTHLY RUN-RATE")
ws_run.cell(row=r_run_total, column=3, value=f"=C{r_infra_sub}+C{r_ops_sub}+C{r_contingency}").number_format = FMT_USD
ws_run.cell(row=r_run_total, column=4, value=f"=D{r_infra_sub}+D{r_ops_sub}+D{r_contingency}").number_format = FMT_USD
style_row(ws_run, r_run_total, 5, font=TOTAL_FONT, fill=total_fill())

wb.defined_names["RUNRATE_MONTHLY"] = DefinedName("RUNRATE_MONTHLY", attr_text=f"'Production Run-Rate'!$C${r_run_total}")
wb.defined_names["RUNRATE_ANNUAL"] = DefinedName("RUNRATE_ANNUAL", attr_text=f"'Production Run-Rate'!$D${r_run_total}")

# V2 deltas
r_v2 = r_run_total + 3
ws_run.cell(row=r_v2-1, column=1, value="— V2 FUTURE ADD-ONS (not in MVP run-rate) —").font = BOLD_FONT
v2_rows = [
    ("V2", "Python calc service (separate plan)", 75, "When baseline calc ships"),
    ("V2", "Claude API — LLM anomaly flagging", 10, "~$0.05/audit × 200 audits"),
    ("V2", "OpenWeatherMap Standard", 40, "For CDD/HDD baseline data"),
    ("V2", "Client portal Static Web App", 9, "When Epic 10 ships"),
]
for cat, item, monthly, notes in v2_rows:
    ws_run.cell(row=r_v2, column=1, value=cat).font = BODY_FONT
    ws_run.cell(row=r_v2, column=2, value=item).font = BODY_FONT
    ws_run.cell(row=r_v2, column=3, value=monthly).number_format = FMT_USD
    ws_run.cell(row=r_v2, column=4, value=f"=C{r_v2}*12").number_format = FMT_USD
    ws_run.cell(row=r_v2, column=5, value=notes).font = BODY_FONT
    for col in range(1, 6):
        ws_run.cell(row=r_v2, column=col).border = border_all
    r_v2 += 1
ws_run.cell(row=r_v2, column=2, value="V2 MONTHLY DELTA")
ws_run.cell(row=r_v2, column=3, value=f"=SUM(C{r_v2-4}:C{r_v2-1})").number_format = FMT_USD
ws_run.cell(row=r_v2, column=4, value=f"=SUM(D{r_v2-4}:D{r_v2-1})").number_format = FMT_USD
style_row(ws_run, r_v2, 5, font=BOLD_FONT, fill=subtotal_fill())

autowidth(ws_run, [14, 36, 14, 14, 40])
ws_run.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════
# Sheet 7: Token Estimates
# ═══════════════════════════════════════════════════════════════
ws_tok = wb.create_sheet("Token Estimates")
ws_tok.sheet_properties.tabColor = "EAB308"
title_row(ws_tok, "Claude Token Estimates (dev + future production)", 9)
tok_headers = ["Phase", "Activity", "Model", "Input tokens", "Output tokens",
               "Cache Read %", "Input $/M", "Output $/M", "Est Cost (USD)"]
set_header(ws_tok, 3, tok_headers)
tok_rows = [
    ("Dev (sunk)", "PRD generation (completed)", "Sonnet 4.6", 500000, 80000, 0.40, 3, 15),
    ("Dev (sunk)", "Architecture (completed)", "Sonnet 4.6", 600000, 100000, 0.50, 3, 15),
    ("Dev (sunk)", "Epics + stories (completed)", "Sonnet 4.6", 800000, 200000, 0.50, 3, 15),
    ("Dev (sunk)", "UX design (completed)", "Sonnet 4.6", 600000, 150000, 0.40, 3, 15),
    ("Dev (MVP)", "Per-story impl session (35 stories)", "Sonnet 4.6", 35*600000, 35*150000, 0.70, 3, 15),
    ("Dev (MVP)", "Per-story review (35 stories)", "Sonnet 4.6", 35*200000, 35*50000, 0.60, 3, 15),
    ("Dev (MVP)", "2nd-terminal coordination premium", "Sonnet 4.6", 1500000, 300000, 0.70, 3, 15),
    ("Dev (MVP)", "Debugging + rework buffer", "Sonnet 4.6", 1500000, 300000, 0.50, 3, 15),
    ("Prod MVP", "LLM anomaly — NOT IN MVP", "Sonnet 4.6", 0, 0, 0, 3, 15),
    ("V2 (future)", "LLM anomaly per audit", "Sonnet 4.6", 17000, 2000, 0.60, 3, 15),
    ("V2 (future)", "200 audits/mo monthly", "Sonnet 4.6", 200*17000, 200*2000, 0.60, 3, 15),
    ("V2 (future)", "If switch to Haiku 4.5 (200/mo)", "Haiku 4.5", 200*17000, 200*2000, 0.60, 1, 5),
    ("V2 (future)", "If switch to Opus 4.7 (200/mo)", "Opus 4.7", 200*17000, 200*2000, 0.60, 15, 75),
]
r = 4
for phase, act, model, inp, out, cache, ipm, opm in tok_rows:
    ws_tok.cell(row=r, column=1, value=phase).font = BODY_FONT
    ws_tok.cell(row=r, column=2, value=act).font = BODY_FONT
    ws_tok.cell(row=r, column=3, value=model).font = BODY_FONT
    ws_tok.cell(row=r, column=4, value=inp).number_format = "#,##0"
    ws_tok.cell(row=r, column=5, value=out).number_format = "#,##0"
    ws_tok.cell(row=r, column=6, value=cache).number_format = FMT_PCT
    ws_tok.cell(row=r, column=7, value=ipm).number_format = FMT_USD
    ws_tok.cell(row=r, column=8, value=opm).number_format = FMT_USD
    # cost = input*(1-cache)*ipm/1M + input*cache*0.1*ipm/1M + output*opm/1M
    ws_tok.cell(row=r, column=9,
        value=f"=((D{r}*(1-F{r})+D{r}*F{r}*0.1)*G{r}+E{r}*H{r})/1000000").number_format = FMT_USD
    for col in range(1, 10):
        ws_tok.cell(row=r, column=col).border = border_all
    r += 1

r_dev_sub = r
ws_tok.cell(row=r_dev_sub, column=2, value="DEV MVP API-equivalent cost (covered by $400 Max sub)")
ws_tok.cell(row=r_dev_sub, column=9, value="=SUM(I8:I11)").number_format = FMT_USD
style_row(ws_tok, r_dev_sub, 9, font=BOLD_FONT, fill=subtotal_fill())

autowidth(ws_tok, [12, 36, 14, 14, 14, 12, 10, 10, 14])
ws_tok.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════
# Sheet 8: Assumptions
# ═══════════════════════════════════════════════════════════════
ws_ass = wb.create_sheet("Assumptions")
ws_ass.sheet_properties.tabColor = SLATE
title_row(ws_ass, "Assumptions (tune these to model scenarios)", 6)
ass_headers = ["ID", "Assumption", "Current Value", "Sensitivity", "If Wrong", "Source"]
set_header(ws_ass, 3, ass_headers)
ass_rows = [
    ("A1", "Your hourly rate", "$20/hour", "High", "Every $5 change = ~$1,150 on project", "User-stated"),
    ("A2", "MVP Claude work-hours (all stories)", "~320 hours", "High", "+50 hrs = +$715 your time", "Per-story estimates"),
    ("A3", "2-terminal parallelism efficiency", "1.4x", "High", "1.2x = +$660; 1.6x = -$410", "Attention/review analysis"),
    ("A4", "Your effective hours (320 / 1.4)", "~230 hours", "High", "Drives labor cost directly", "Derived"),
    ("A5", "Project duration — MVP build", "6 weeks full-time", "High", "+1 week = +$240 carry cost", "40 hrs/wk × 5.75 weeks"),
    ("A6", "Calendar duration (build + launch)", "2 months", "Medium", "Launch month accounts for prod Azure", "Plan assumption"),
    ("A7", "Claude plan", "Max 20x ($200/mo)", "High", "Pro = rate-limit hell with 2 terminals", "Max 20x required"),
    ("A8", "MVP scope", "Refrigeration + PDF email", "High", "HVAC/Lighting added back = +150 hrs", "Aggressive cut"),
    ("A9", "Hired developer", "No", "Low", "Post-MVP contractor possible $60/hr × 50hr = +$3k", "LLM-first solo build"),
    ("A10", "Audits per month at launch", "<50 (beta)", "Medium", ">100 forces tier upgrades", "Internal use initially"),
    ("A11", "Azure region", "Canada Central", "Low", "+/- 10% egress cost", "Locked"),
    ("A12", "Reserved instances", "No (pay-as-you-go)", "Medium", "1yr reserved = -30% infra", "Revisit 6 months post-launch"),
    ("A13", "Security audit pre-MVP-beta", "Skipped", "High", "Blocker for enterprise clients", "Internal MVP only"),
    ("A14", "Legal lawyer review", "Skipped (templates)", "Medium", "Needed before billing clients", "Defer to V2"),
    ("A15", "LLM anomaly flagging in MVP", "Deferred to V2", "Medium", "Save 24 hrs + $10/mo prod API", "User-accepted cut"),
    ("A16", "Client portal in MVP", "Deferred to V2", "Medium", "Email PDF replaces it", "User-accepted cut"),
    ("A17", "Python calc service", "Consolidated into Node", "Low", "Simpler deployment", "MVP simplification"),
    ("A18", "SendGrid", "Free tier", "Low", ">100 emails/day = +$20/mo", "MVP beta volume"),
    ("A19", "Weather API", "Not needed in MVP", "Low", "Needed when baseline calc lands", "Deferred"),
    ("A20", "Contingency buffer", "15%", "Medium", "Risky = 25%", "Standard PM practice"),
    ("A21", "Epic 0 spikes completion", "Before feature work", "High", "RLS failure = rebuild security model", "Non-negotiable gate"),
    ("A22", "Design tokens minimal (Tailwind)", "Yes", "Low", "Save 6-8 hrs vs full system", "MVP simplification"),
]
r = 4
for aid, ass, val, sens, wrong, src in ass_rows:
    ws_ass.cell(row=r, column=1, value=aid).font = BOLD_FONT
    ws_ass.cell(row=r, column=2, value=ass).font = BODY_FONT
    ws_ass.cell(row=r, column=3, value=val).font = BODY_FONT
    sens_cell = ws_ass.cell(row=r, column=4, value=sens)
    sens_cell.font = BODY_FONT
    sens_cell.alignment = ALIGN_CENTER
    if sens == "High":
        sens_cell.fill = PatternFill("solid", fgColor="FECACA")
    elif sens == "Medium":
        sens_cell.fill = PatternFill("solid", fgColor="FED7AA")
    else:
        sens_cell.fill = PatternFill("solid", fgColor="DCFCE7")
    ws_ass.cell(row=r, column=5, value=wrong).font = BODY_FONT
    ws_ass.cell(row=r, column=6, value=src).font = BODY_FONT
    for col in range(1, 7):
        ws_ass.cell(row=r, column=col).border = border_all
        ws_ass.cell(row=r, column=col).alignment = Alignment(
            horizontal="center" if col in (1, 4) else "left",
            vertical="center", wrap_text=True)
    r += 1
autowidth(ws_ass, [8, 42, 22, 13, 50, 22])
ws_ass.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════
# Sheet 9: Actuals Log
# ═══════════════════════════════════════════════════════════════
ws_act = wb.create_sheet("Actuals Log")
ws_act.sheet_properties.tabColor = "DC2626"
title_row(ws_act, "Actuals Log — record real spend as project progresses", 10)
act_headers = ["Date", "Category", "Subcategory", "Vendor", "Description",
               "Amount (USD)", "Payment Method", "Invoice #", "Linked To", "Notes"]
set_header(ws_act, 3, act_headers)
starter_rows = [
    ("2026-04-19", "Labor", "Planning", "Self", "PRD draft and architecture review", None),
    ("2026-04-20", "Labor", "Story execution", "Self", "Story 0-1 Turborepo scaffold", None),
    ("2026-04-22", "Labor", "Story execution", "Self", "Story 0-2 Azure infrastructure provisioning", None),
    ("2026-04-24", "Labor", "Story execution", "Self", "Story 0-3 DB schema + RLS (partial)", None),
    (None, "Subscriptions", "Claude Max 20x", "Anthropic", "Monthly subscription", 200),
    (None, "Subscriptions", "GitHub Free", "GitHub", "No charge", 0),
    (None, "Infrastructure", "Azure dev environment", "Microsoft", "First bill ~30 days after provisioning", None),
    (None, "One-time", "Domain", "TBD registrar", "starenergy-cems.com (or similar)", 15),
    (None, "One-time", "Legal templates", "Termly", "T&Cs + Privacy", 50),
]
r = 4
for row in starter_rows:
    for col, v in enumerate(row, start=1):
        cell = ws_act.cell(row=r, column=col, value=v)
        cell.font = BODY_FONT
        cell.border = border_all
    amt = ws_act.cell(row=r, column=6)
    amt.number_format = FMT_USD
    r += 1
# Leave blank rows for future entries
for _ in range(30):
    for col in range(1, 11):
        c = ws_act.cell(row=r, column=col)
        c.border = border_all
        if col == 6:
            c.number_format = FMT_USD
    r += 1
r_act_total = r + 1
ws_act.cell(row=r_act_total, column=5, value="TOTAL ACTUAL SPEND (cash)")
ws_act.cell(row=r_act_total, column=6, value=f"=SUM(F4:F{r-1})").number_format = FMT_USD
style_row(ws_act, r_act_total, 10, font=TOTAL_FONT, fill=total_fill())

autowidth(ws_act, [12, 14, 18, 18, 40, 14, 16, 14, 14, 30])
ws_act.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════
# Sheet 1: Dashboard — build last so all refs resolve
# ═══════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("Dashboard", 0)
ws_dash.sheet_properties.tabColor = NAVY

ws_dash.cell(row=1, column=1, value="Star Energy CEMS — MVP Cost Tracker").font = Font(
    name="Calibri", size=18, bold=True, color=NAVY)
ws_dash.merge_cells("A1:F1")
ws_dash.cell(row=2, column=1, value="2-terminal parallel build • Claude Max 20x • $20/hr • 6-week MVP").font = Font(
    name="Calibri", size=11, italic=True, color=SLATE)
ws_dash.merge_cells("A2:F2")
ws_dash.row_dimensions[1].height = 30
ws_dash.row_dimensions[2].height = 20

set_header(ws_dash, 4, ["Section", "Metric", "Estimated (USD)", "Actual (USD)", "Variance (USD)", "% of Build"])

r = 5
def dash_row(section, metric, est_formula, notes_row=False, highlight=False):
    global r
    ws_dash.cell(row=r, column=1, value=section).font = BODY_FONT
    ws_dash.cell(row=r, column=2, value=metric).font = BOLD_FONT if highlight else BODY_FONT
    cell = ws_dash.cell(row=r, column=3, value=est_formula)
    cell.number_format = FMT_USD
    ws_dash.cell(row=r, column=4).number_format = FMT_USD
    ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
    ws_dash.cell(row=r, column=6, value=f"=C{r}/$C$BUILD_TOTAL_ROW").number_format = FMT_PCT
    for col in range(1, 7):
        ws_dash.cell(row=r, column=col).border = border_all
        if highlight:
            ws_dash.cell(row=r, column=col).fill = subtotal_fill()
            ws_dash.cell(row=r, column=col).font = BOLD_FONT
    r += 1

# Labor
ws_dash.cell(row=r, column=1, value="LABOR").font = Font(name="Calibri", size=11, bold=True, color=TEAL)
ws_dash.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT)
ws_dash.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws_dash.cell(row=r, column=1).alignment = ALIGN_LEFT
for col in range(1, 7):
    ws_dash.cell(row=r, column=col).border = border_all
r += 1

labor_row_start = r
ws_dash.cell(row=r, column=1, value="Labor").font = BODY_FONT
ws_dash.cell(row=r, column=2, value="Claude work-hours (MVP in-scope)").font = BODY_FONT
ws_dash.cell(row=r, column=3, value="=LABOR_HOURS_TOTAL").number_format = FMT_HRS
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
r += 1
ws_dash.cell(row=r, column=1, value="Labor").font = BODY_FONT
ws_dash.cell(row=r, column=2, value="Your effective hours (÷ 1.4)").font = BODY_FONT
ws_dash.cell(row=r, column=3, value="=LABOR_EFFECTIVE_HOURS").number_format = FMT_HRS
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
r += 1
ws_dash.cell(row=r, column=1, value="Labor").font = BODY_FONT
ws_dash.cell(row=r, column=2, value="TOTAL LABOR COST @ $20/hr").font = BOLD_FONT
c = ws_dash.cell(row=r, column=3, value="=LABOR_COST")
c.number_format = FMT_USD
c.font = BOLD_FONT
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7):
    ws_dash.cell(row=r, column=col).border = border_all
    ws_dash.cell(row=r, column=col).fill = subtotal_fill()
labor_cost_row = r
r += 2

# Claude
ws_dash.cell(row=r, column=1, value="CLAUDE & TOOLS").font = Font(name="Calibri", size=11, bold=True, color=AMBER)
ws_dash.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT)
ws_dash.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
r += 1
ws_dash.cell(row=r, column=1, value="Claude").font = BODY_FONT
ws_dash.cell(row=r, column=2, value="Claude Max 20x × 2 months").font = BODY_FONT
ws_dash.cell(row=r, column=3, value="=SUBS_REQUIRED_TOTAL").number_format = FMT_USD
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
claude_row = r
r += 1
ws_dash.cell(row=r, column=1, value="Claude").font = BODY_FONT
ws_dash.cell(row=r, column=2, value="Production LLM API (MVP — zero)").font = BODY_FONT
c = ws_dash.cell(row=r, column=3, value=0); c.number_format = FMT_USD
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
r += 2

# Infra
ws_dash.cell(row=r, column=1, value="INFRASTRUCTURE").font = Font(name="Calibri", size=11, bold=True, color=AMBER)
ws_dash.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT)
ws_dash.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
r += 1
ws_dash.cell(row=r, column=1, value="Infra").font = BODY_FONT
ws_dash.cell(row=r, column=2, value="Azure dev (2 months minimum tiers)").font = BODY_FONT
ws_dash.cell(row=r, column=3, value="=INFRA_DEV_TOTAL").number_format = FMT_USD
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
infra_dev_row = r
r += 1
ws_dash.cell(row=r, column=1, value="Infra").font = BODY_FONT
ws_dash.cell(row=r, column=2, value="Azure prod (launch month)").font = BODY_FONT
ws_dash.cell(row=r, column=3, value="=INFRA_PROD_TOTAL").number_format = FMT_USD
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
infra_prod_row = r
r += 1
ws_dash.cell(row=r, column=1, value="Infra").font = BODY_FONT
ws_dash.cell(row=r, column=2, value="Third-party services (2 months)").font = BODY_FONT
ws_dash.cell(row=r, column=3, value="=INFRA_3P_TOTAL").number_format = FMT_USD
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
infra_3p_row = r
r += 2

# One-time
ws_dash.cell(row=r, column=1, value="ONE-TIME").font = Font(name="Calibri", size=11, bold=True, color=AMBER)
ws_dash.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT)
ws_dash.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
r += 1
ws_dash.cell(row=r, column=1, value="Setup").font = BODY_FONT
ws_dash.cell(row=r, column=2, value="Domain + legal templates + training").font = BODY_FONT
ws_dash.cell(row=r, column=3, value="=ONETIME_TOTAL").number_format = FMT_USD
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
onetime_row = r
r += 2

# Totals
r_subtotal = r
ws_dash.cell(row=r, column=2, value="SUBTOTAL (before contingency)").font = BOLD_FONT
ws_dash.cell(row=r, column=3,
    value=f"=C{labor_cost_row}+C{claude_row}+C{infra_dev_row}+C{infra_prod_row}+C{infra_3p_row}+C{onetime_row}").number_format = FMT_USD
ws_dash.cell(row=r, column=4, value=f"=D{labor_cost_row}+D{claude_row}+D{infra_dev_row}+D{infra_prod_row}+D{infra_3p_row}+D{onetime_row}").number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7):
    ws_dash.cell(row=r, column=col).border = border_all
    ws_dash.cell(row=r, column=col).fill = subtotal_fill()
    ws_dash.cell(row=r, column=col).font = BOLD_FONT
r += 1

r_contingency_d = r
ws_dash.cell(row=r, column=2, value="Contingency (15%)").font = BODY_FONT
ws_dash.cell(row=r, column=3, value=f"=C{r_subtotal}*0.15").number_format = FMT_USD
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
r += 1

r_build_total = r
ws_dash.cell(row=r, column=2, value="TOTAL TO LAUNCH MVP").font = TOTAL_FONT
ws_dash.cell(row=r, column=3, value=f"=C{r_subtotal}+C{r_contingency_d}").number_format = FMT_USD
ws_dash.cell(row=r, column=4, value=f"=D{r_subtotal}+D{r_contingency_d}").number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7):
    ws_dash.cell(row=r, column=col).border = border_all
    ws_dash.cell(row=r, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws_dash.cell(row=r, column=col).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
r += 2

# Now that r_build_total is known, replace the BUILD_TOTAL_ROW placeholder
for row_cells in ws_dash.iter_rows(min_row=5, max_row=r_build_total, min_col=6, max_col=6):
    for cell in row_cells:
        if cell.value and isinstance(cell.value, str) and "BUILD_TOTAL_ROW" in cell.value:
            cell.value = cell.value.replace("$C$BUILD_TOTAL_ROW", f"$C${r_build_total}")

# Cash-only
r_cash = r
ws_dash.cell(row=r, column=1, value="CASH OUT-OF-POCKET").font = Font(name="Calibri", size=11, bold=True, color=TEAL)
ws_dash.cell(row=r, column=2, value="Cash only (excludes your sweat-equity labor)").font = BOLD_FONT
ws_dash.cell(row=r, column=3,
    value=f"=C{claude_row}+C{infra_dev_row}+C{infra_prod_row}+C{infra_3p_row}+C{onetime_row}+(C{r_contingency_d}*0.2)").number_format = FMT_USD
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7):
    ws_dash.cell(row=r, column=col).border = border_all
    ws_dash.cell(row=r, column=col).fill = PatternFill("solid", fgColor="D1FAE5")
    ws_dash.cell(row=r, column=col).font = BOLD_FONT
r += 2

# Year 1 run-rate
r_year1 = r
ws_dash.cell(row=r, column=1, value="YEAR 1 OPS").font = Font(name="Calibri", size=11, bold=True, color="7C3AED")
ws_dash.cell(row=r, column=2, value="Year 1 production run-rate (post-launch)").font = BODY_FONT
ws_dash.cell(row=r, column=3, value="=RUNRATE_ANNUAL").number_format = FMT_USD
ws_dash.cell(row=r, column=4).number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7): ws_dash.cell(row=r, column=col).border = border_all
r += 1

r_grand = r
ws_dash.cell(row=r, column=2, value="GRAND TOTAL YEAR 1 (build + run)").font = TOTAL_FONT
ws_dash.cell(row=r, column=3, value=f"=C{r_build_total}+C{r_year1}").number_format = FMT_USD
ws_dash.cell(row=r, column=4, value=f"=D{r_build_total}+D{r_year1}").number_format = FMT_USD
ws_dash.cell(row=r, column=5, value=f"=D{r}-C{r}").number_format = FMT_USD
for col in range(1, 7):
    ws_dash.cell(row=r, column=col).border = border_all
    ws_dash.cell(row=r, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws_dash.cell(row=r, column=col).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
r += 3

# Scenarios
ws_dash.cell(row=r, column=1, value="SCENARIOS").font = Font(name="Calibri", size=12, bold=True, color=NAVY)
r += 1
scen_headers = ["Scenario", "Description", "Total (USD)"]
for i, h in enumerate(scen_headers, 1):
    c = ws_dash.cell(row=r, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = header_fill()
    c.alignment = ALIGN_CENTER
    c.border = border_all
r += 1

scenarios = [
    ("Lean", "Scope held, no contingency burned", f"=C{r_subtotal}"),
    ("Realistic", "This dashboard (default)", f"=C{r_build_total}"),
    ("Slow", "Part-time 25hr/wk → 10 weeks calendar", f"=(LABOR_HOURS_TOTAL*1.7*20)+(C{claude_row}*2.5)+(C{infra_dev_row}*2.5)+C{infra_prod_row}+(C{infra_3p_row}*2.5)+C{onetime_row}"),
    ("Scope creep", "HVAC + Lighting + Envelope added back", f"=(LABOR_HOURS_TOTAL*1.5*20)+(C{claude_row}*1.5)+(C{infra_dev_row}*1.5)+C{infra_prod_row}+(C{infra_3p_row}*1.5)+C{onetime_row}"),
]
for name, desc, formula in scenarios:
    ws_dash.cell(row=r, column=1, value=name).font = BOLD_FONT
    ws_dash.cell(row=r, column=2, value=desc).font = BODY_FONT
    cell = ws_dash.cell(row=r, column=3, value=formula)
    cell.number_format = FMT_USD
    for col in range(1, 4):
        ws_dash.cell(row=r, column=col).border = border_all
    r += 1

r += 2
ws_dash.cell(row=r, column=1, value="NEXT ACTIONS").font = Font(name="Calibri", size=12, bold=True, color=NAVY)
r += 1
actions = [
    "1. Finalize Epic 0 spikes (RLS, Puppeteer, Redis) before forking terminals",
    "2. Confirm Terminal A/B file-system split (see 10-terminal-execution-plan.md)",
    "3. Log actual hours on 'Actuals Log' tab daily",
    "4. Review assumptions weekly; update values as reality diverges",
    "5. If contingency burns past 50% by Week 4 → cut more scope, not overrun",
]
for a in actions:
    ws_dash.cell(row=r, column=1, value=a).font = BODY_FONT
    ws_dash.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

autowidth(ws_dash, [16, 44, 18, 18, 16, 12])
ws_dash.freeze_panes = "A5"

# ═══════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════
out = Path(__file__).parent / "Star-Energy-CEMS-Cost-Tracker.xlsx"
wb.save(out)
print(f"Built: {out}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
